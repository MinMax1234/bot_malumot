import os
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

COLUMNS = [
    ("telegram_id",       "Telegram ID"),
    ("full_name",         "To'liq ism"),
    ("position",          "Lavozim"),
    ("degree",            "Ta'lim darajasi"),
    ("experience_years",  "Tajriba (yil)"),
    ("university",        "Oliy ta'lim muassasasi"),
    ("specialization",    "Mutaxassislik"),
    ("subjects",          "O'qitiladigan fanlar"),
    ("has_ielts",         "IELTS bormi"),
    ("ielts_score",       "IELTS ball"),
    ("has_cefr",          "CEFR bormi"),
    ("cefr_level",        "CEFR daraja"),
    ("has_national_cert", "Milliy sertifikat"),
    ("bio",               "Bio"),
    ("photo_path",        "Rasm fayli"),
    ("awards",            "Mukofotlar"),
    ("created_at",        "Ro'yxatdan o'tgan sana"),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
EVEN_FILL     = PatternFill("solid", fgColor="D6E4F0")
THIN          = Side(style="thin", color="B0B0B0")
CELL_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_excel(rows: list[dict], output_path: str | None = None) -> str:
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"ustozlar_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ustozlar"

    # Header row
    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, record in enumerate(rows, start=2):
        fill = EVEN_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, (field, _) in enumerate(COLUMNS, start=1):
            value = record.get(field)
            if field in ("has_ielts", "has_cefr", "has_national_cert"):
                value = "Ha" if value else "Yo'q"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
            if fill.fill_type:
                cell.fill = fill

    # Column widths
    widths = [14, 24, 22, 20, 12, 28, 22, 26, 12, 10, 10, 10, 16, 40, 22, 40, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path


def export_photos_zip(rows: list[dict], output_path: str | None = None) -> tuple[str, int]:
    """Barcha ustozlar rasmlarini ZIP arxivga yig'adi.

    Returns: (zip_fayl_yoli, rasm_soni)
    """
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"rasmlar_{ts}.zip"

    count = 0
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for record in rows:
            path = record.get("photo_path")
            if path and os.path.exists(path):
                name = record.get("full_name") or str(record.get("telegram_id", "unknown"))
                # Fayl nomini xavfsiz qilamiz
                safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in name).strip()
                ext = os.path.splitext(path)[1] or ".jpg"
                zf.write(path, arcname=f"{safe_name}{ext}")
                count += 1

    return output_path, count
